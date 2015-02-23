# -*- coding: utf-8 -*-

# EXCEL read write librarys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import numpy as np

import os,copy
import warnings


class DeioHelper:

    wb = 0    # active workbook object
    timespan =[]   # eg["2008","2009","2010"]
    dir_output =""
    sectors=[]

    def __init__(self,filepath,name_cntry,timespan,dir_output,sectors):
        warnings.filterwarnings("ignore")
        self.timespan=timespan
        self.dir_output=dir_output
        self.sectors=sectors

        ##print("output")
        ##print sectors.__len__()

        #...load workbook
        print("Load Data from: %s " %name_cntry),

        try:
            self.wb =  load_workbook(filename = filepath,data_only = True,use_iterators = True)
            print("...OK")
        except:
            print("+++++++++ERROR AT LOADING FILE ++++++++++++")

        warnings.filterwarnings("default")



##-----------------------------------------------------------------------
    def get_data(self,sheetname,data_range):
        # returns data from workbooks area in 2D style [row][col]
        #sheetname  =  ["use","pp"] data_range = "D8:BY81"

        target = []
        ws1  =  self.wb.get_sheet_by_name(name  =  sheetname) # eg. sup08pp

        for i,row in enumerate(ws1.iter_rows(data_range)):   # row
            target.append([])
            for j,cell in enumerate(row):                            #column
                target[i].append(cell.value)

        return target



#-------------------- 2 Dimensional MATRIX OPERATIONS---------------------

    def gen_2d_matrix(self,rows,cols):  # GENERATE 0 Matrix
        target = []
        for i in range(rows):
            target.append([])
            for j in range(cols):
                target[i].append(0)
        return target

##-----------------------------------------------------------------------
    def transpose_2d(self,data):        # TRANSPOSE
        rows =  data.__len__()
        cols =  data[0].__len__()

        target = self.gen_2d_matrix(cols,rows)

        for j,rw in enumerate(data):
            for k,cl in enumerate(rw):
                target[k][j] = data[j][k]
        return target

##-----------------------------------------------------------------------
    def add_matrix_2d(self,data1,data2):     # ADD two MATRIXes
        target = list(data1)
        for i,rows in enumerate(data1):
            for j,cl in enumerate(rows):
                target[i][j]=data[i][j]+data[i][j]
    #return np.add(a, b)

##-----------------------------------------------------------------------
    def extract_rows_2d(self,data, rows_id):   # EXTRACT ROWS from 2d eg. 2d(data[row][col]) ; rows_id=[10,11,12]


        anz_cols = len(data[0])       # number of columns stays
        try:
            anz_new_rows = len(rows_id)   # number of rows to extract
        except:
            print("fsadfasdfsa %s" %rows_id)


        target=self.gen_2d_matrix(anz_new_rows,anz_cols)  # make target matrix

        # FILL DATA
        for i,rows in enumerate(data):
            for j,cell in enumerate(rows):
                if i in rows_id:           #if row "i" is in rows_id [2,4,5]
                    k=rows_id.index(i) # k is the row index within the new matrix
                    target[k][j]=cell


        return target

##-----------------------------------------------------------------------
    def extract_cols_2d(self,data, cols_id):    # EXTRACT COLS
        anz_new_cols = len(cols_id)
        anz_rows = len(data)

        target=self.gen_2d_matrix(anz_rows,anz_new_cols)

        for i,rows in enumerate(data):
            for j,cell in enumerate(rows):
                if j in cols_id:
                    k=cols_id.index(j) # k is the col index within the new matrix
                    target[i][k]=cell

        return target

##-----------------------------------------------------------------------
    def delete_rows_2d(self,data):



        pass
##-----------------------------------------------------------------------
    def delete_cols_2d(self):
        pass
##-----------------------------------------------------------------------
    def share_over_rows_2d(self,data):
        ##print("len %s %s " %(data.__len__(),data[0].__len__()))

        target = self.gen_2d_matrix(data.__len__(),data[0].__len__())  # AGGREG ROWS

        rowsum=self.aggreg_rows_2d(data)

        for i,row in enumerate(data):
            for j, cl in enumerate(row):
                if rowsum[0][j]!=0:
                    target[i][j]=data[i][j]/rowsum[0][j]
                else:
                    target[i][j]=0
        return target


##-----------------------------------------------------------------------
    def share_over_cols_2d(self,data):
        target = self.gen_2d_matrix(data.__len__(),data[0].__len__())
        colsum=self.aggreg_cols_2d(data)

        for i,row in enumerate(data):
            for j, cl in enumerate(row):
                if colsum[i][0]!=0:
                    target[i][j]=data[i][j]/colsum[i][0]
                else:
                    target[i][j]=0
        return target

##-----------------------------------------------------------------------
    def divide_by_row_2d(self,data,rowdata):  # a nxn matrix elementwise divide by a 1xn matrix (for SED,SEM / EN)
        target = list(data) # make new target

        for i,rw_data in enumerate(data):
            for j,cell in enumerate(rw_data):
                if rowdata[0][j]!=0:
                    target[i][j] = data[i][j]/rowdata[0][j]
                else:
                    target[i][j] =0


        return target


##-----------------------------------------------------------------------
    def aggreg_rows_2d(self,data):   # returns  nx1 vector [n][0]
        target = self.gen_2d_matrix(1,data[0].__len__())  # AGGREG ROWS

        for i,rw_data in enumerate(data):
            for j,cell in enumerate(rw_data):
                target[0][j]+= cell

        return target

##-----------------------------------------------------------------------
    def aggreg_cols_2d(self,data):  # returns  nx1 vector [n][0]
        target = self.gen_2d_matrix(data.__len__(),1)  # AGGREG COLS

        for i,rw_data in enumerate(data):
            for j,cell in enumerate(rw_data):
                target[i][0]+= cell

        return target

##-----------------------------------------------------------------------
    def make_rows_to_1column_2d(self,data): # make from 5x10 -> 100x1  # make rows to 1 column

        ##print("len %s %s " %(data.__len__(),data[0].__len__()))

        rowdim =data[0].__len__()

        target = self.gen_2d_matrix(data.__len__()*data[0].__len__(),1)

        for i,row in enumerate(data):
            for j,cell in enumerate(row):
                target[rowdim*i+j][0]= data[i][j]

        return target


    def make_cols_to_1column_2d(self,data): # make from 5x10 -> 100x1  # make columns to 1 column

        coldim =data.__len__()

        target = self.gen_2d_matrix(data.__len__()*data[0].__len__(),1)

        for i,row in enumerate(data):
            for j,cell in enumerate(row):
                target[coldim*i+j][0]= data[i][j]

        return target

#----------------------------------------------------------------
#---------------------------- IPTS FUNCTIONS --------------------
#----------------------------------------------------------------

    def IPTS_aggreg_rows(self,data):            # AGGREG ROWS

        target = []

        for i,yr_data in enumerate(data):
            target.append(self.aggreg_rows_2d(yr_data))

        return target

##-----------------------------------------------------------------------
    def IPTS_aggreg_cols(self,data):            # AGGREG COLS
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.aggreg_cols_2d(yr_data))

        return target

##-----------------------------------------------------------------------
    def IPTS_share_over_rows(self,data):            # share ROWS

        target = []

        for i,yr_data in enumerate(data):
            target.append(self.share_over_rows_2d(yr_data))

        return target


##-----------------------------------------------------------------------
    def IPTS_share_over_cols(self,data):            # share COLS
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.share_over_cols_2d(yr_data))

        return target

##-----------------------------------------------------------------------
    def IPTS_divide_by_row(self,data,rowdata):            # share ROWS

        target = []

        for i,yr_data in enumerate(data):
            target.append(self.divide_by_row_2d(yr_data,rowdata[i]))

        return target



##-----------------------------------------------------------------------
    def IPTS_transpose(self,data):  # Transposes IPTS-3Dim
        orig_yrs =  data.__len__()   # nr years
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.transpose_2d(yr_data))

        return target


    def IPTS_make_rows_to_1column(self,data):  # converts 2d -> 1d
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.make_rows_to_1column_2d(yr_data))

        return target

#--------------------------------------------
##    def IPTS_make_cols_to_1column(self,data):  # converts 2d -> 1d
##        target = []
##
##        for i,yr_data in enumerate(data):
##            target.append(self.make_cols_to_1column_2d(yr_data))
##
##        return target

#--------------------------------------------

    def IPTS_extract_rows(self,data,rows_id):
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.extract_rows_2d(yr_data,rows_id))

        return target
##-----------------------------------------------------------------------
    def IPTS_extract_cols(self,data,cols_id):
        target = []

        for i,yr_data in enumerate(data):
            target.append(self.extract_cols_2d(yr_data,cols_id))

        return target




##-----------------------------------------------------------------------
    def IPTS_gen_matrix(self,yrs,rows,cols):  #gens 3d 0-Matrix in IPTS stlye
        target = []
        for i in range(yrs):
            target.append([])
            for j in range(rows):
                target[i].append([])
                for k in range(cols):
                    target[i][j].append(0)
        return target


##-----------------------------------------------------------------------
    def IPTS_sub_elem(self,a,b):  # c= a - b
        c = list(a) # create new
        for yr,year in enumerate(c):
            for i,row in enumerate(year):
                for j,cl in enumerate(row):
                    c[yr][i][j]=a[yr][i][j]-b[yr][i][j]
        return c


##-----------------------------------------------------------------------
    def IPTS_add_elem(self,a,b):  # c= a - b
        c = list(a) # create new
        for yr,year in enumerate(c):
            for i,row in enumerate(year):
                for j,cl in enumerate(row):
                    c[yr][i][j]=a[yr][i][j]+b[yr][i][j]
        return c


##-----------------------------------------------------------------------
    def IPTS_div_elem(self,a,b):  # c= a / b
        c = list(a) # create new
        for yr,year in enumerate(c):
            for i,row in enumerate(year):
                for j,cl in enumerate(row):
                    if b[yr][i][j]!=0:
                        c[yr][i][j]=a[yr][i][j]/b[yr][i][j]  # SEctor 97  is 0
                    else:
                         c[yr][i][j]=0

        return c

#--------------------------------------------------------

    def write_simple(self,data,varname="test",info="-"):   # write 3dim Data

        wb  =  Workbook()            # generate separate file for variable
        ws  =  wb.create_sheet(0)
        ws.title= varname

        offset=1
        ##rowdimension =len(data[0])   # anzahl der reihen
        ##sector_count = len(self.sectors)   # anzahl der sektoren


        ws.cell(row=1,column=1).value=info   # INFO

        for yr,yr_data in enumerate(data):
            ws.cell(row=1,column=yr+1+offset).value = self.timespan[yr]  #LABEL col (yrs)

            for i,row in enumerate(yr_data):
                ## IN ARBEIT ws.cell(row=i+1+offset,column=1).value=vlabels[i]   # LABEL row (vars)

                # FILL Data
                for j,cl in enumerate(row):
                    ws.cell(row  =  i+1+offset, column  = j+1+offset+yr).value  =  cl

        # SAVE
        wb.save(os.path.join(self.dir_output,varname+".xlsx"))




    def aggregate_data_of_a_directory(self,fname):   # write 3dim Data

        print("Aggregating all data..."),

        targetname = os.path.join(self.dir_output,fname+".xlsx")
        if fname+".xlsx" in os.listdir(self.dir_output):
            os.remove(targetname)

        wb_target  =  Workbook()

        # iterate through files
        for nr,i in enumerate(os.listdir(self.dir_output)):           # iterate over all xlsx files
            if i.endswith(".xlsx"):

                wsname = i.split(".")[0]

                wb_src =  load_workbook(filename = os.path.join(self.dir_output,i))
                ws_src = wb_src.get_sheet_by_name(name = wsname)            # load the first sheet

                rd = ws_src.get_highest_row()
                cd = ws_src.get_highest_column()       # get sheet dimensions

                ws_target = wb_target.create_sheet(0)  # create new sheet

                for i in range(rd): #row
                    for j in range(cd): #col
                        ws_target.cell(row = i+1, column = j+1).value =  ws_src.cell(row = i+1, column = j+1).value

                ws_target.title = wsname


        wb_target.save(targetname)
        print("...Success!")

#----------------------- CHECK IPTS METHODS-----------------

    def check_rows(self):
        pass
# check row sums, col sums, Supply = Demand  pp = bp+ttm,tls usw...




##    def write_to_xlsx(self,name,data,info,is2dim = False):
##
##        print("START filling information and data %s..." %name),
##
##        ##wb =  load_workbook(filename = os.path.join(self.dir_output,"test.xlsx"))
##        wb  =  Workbook()
##        ws  =  wb.create_sheet(0)
##        ws.setitle = name
##
##        off_rw = 1
##        off_cl = 1
##
##        sec_set = self.sectors
##
##        # fill in information
##        ws.cell(row  = 1, column = 1).value  =  info
##
##        for i,yr in enumerate(self.timespan):       #years
##            ws.cell(row  =  1 , column  =  i+1+off_cl).value = "20"+yr
##
##        if twodim == False:
##            for j,sec in enumerate(sec_set):       #data names
##                ws.cell(row = 1+off_rw+j , column  =  1).value  =  name+sec  # eg QN02
##
##        if twodim == True:
##            anz = sec_set.__len__() #number of sectors
##            for j,sec_r in enumerate(sec_set):       #data row
##                for k,sec_c in enumerate(sec_set):  #data column
##                    ws.cell(row = 1+off_rw+(j*anz+k) , column  =  1).value  =  name+sec_r+"_"+sec_c  # eg D01_01, D02_01, D05_01...
##
##        # fill in data
##        for i,col in enumerate(data):
##            for j,cl in enumerate(col):
##                ws.cell(row  =  j+1+off_rw, column  =  i+1+off_cl).value  =  cl
##
##
##        wb.save(os.path.join(self.dir_output,name+".xlsx"))
##        print("...OK")



##    def sum_over_cols(self,data):
##
##        target = self.gen_matrix(data.__len__(),data[0].__len__(),1)
##
##        for i,yr_data in enumerate(data):
##            for j,rw_data in enumerate(yr_data):
##                tmp = 0
##                for k,cell in enumerate(rw_data):
##                    tmp+= cell
##
##                target[i][j][0] = tmp
##        return target



##
##
##    def transpose(self,data):  # Transposes IPTS-3Dim
##        orig_yrs =  data.__len__()   # nr years
##        orig_rows =  data[0].__len__() # nr rows
##        orig_cols =  data[0][0].__len__() # nr rows
##        target = self.gen_matrix(orig_yrs,orig_cols,orig_rows)
##
##        # change rows and columns
##        for i,yr in enumerate(data):
##            for j,rw in enumerate(yr):
##                for k,cl in enumerate(rw):
##                    target[i][k][j] = data[i][j][k]
##        return target


