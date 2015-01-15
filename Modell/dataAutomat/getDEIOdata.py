# -*- coding: utf-8 -*-

# EXCEL read write librarys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import numpy as np

import os

class GetSUT_IPTS:
    filepath=""                         # full path of SUT
    wb=0                                # active workbook object
    timespan=""                         # eg["08","09","10"]
    sectors = ['01', '02', '03', '05', '08', '10', '11', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '35', '36', '37', '41', '42', '43', '45', '46', '47', '49', '50', '51', '52', '53', '55', '58', '59', '60', '61', '62', '64', '65', '66', '68', '69', '70', '71', '72', '73', '74', '77', '78', '79', '80', '84', '85', '86', '87', '90', '91', '92', '93', '94', '95', '96', '97']
    dir_output='K:\\Projekt_DEIO\\Data\\_InputData'


    def __init__(self,path_sut,name_cntry,timespan):

        self.filepath = os.path.join(path_sut,name_cntry)
        self.timespan=timespan

        #...load workbook
        print("Loading Datafile %s " %name_cntry)
        self.wb= load_workbook(filename=self.filepath,data_only=True,use_iterators=True)



    def get_array(self,su_cat,prc_cat,dataarea):  #eg. "sup","pp","A2:D3"
    #returns a area as an array for each year [ [x,x,x,x],[x,x,x,x,x],[x,x,x,x,x] ]
        print("Start extract data..."),
        target=[]

        for i,yr in enumerate(self.timespan):
            target.append([])  # new column for each year

            ws1 = self.wb.get_sheet_by_name(name = su_cat + yr + prc_cat) # eg. sup08pp

            for row in ws1.iter_rows(dataarea):
                for cell in row:
                    target[i].append(cell.value)

        print("...OK")
        return target


    def write_array(self,name,data,info):
        print("START filling information and data ..."),


        ##wb= load_workbook(filename=os.path.join(self.dir_output,"test.xlsx"))
        wb = Workbook()
        ws = wb.create_sheet(0)
        ws.title=name

        off_rw=1
        off_cl=1


        # fill in information
        ws.cell(row =1, column=1).value = info

        for i,yr in enumerate(self.timespan):       #years
            ws.cell(row = 1 , column = i+1+off_cl).value="20"+yr

        for j,sec in enumerate(self.sectors):       #data names
            ws.cell(row=1+off_rw+j , column = 1).value = name+sec  # eg QN02

        # fill in data
        for i,col in enumerate(data):
            for j,cl in enumerate(col):
                ws.cell(row = j+1+off_rw, column = i+1+off_cl).value = cl



        wb.save(os.path.join(self.dir_output,name+".xlsx"))
        print("...OK")


    def add_arrays(self,a,b):
        return np.add(a, b)
        #for yr in timespan:
         #   for

    def div_arrays(self,a,b):
        return np.true_divide(a,b)


