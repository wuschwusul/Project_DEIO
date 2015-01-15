# -*- coding: utf-8 -*-


#izi_xlHandler


# openwb, openworkbook(path of xls,xlsx)


# EXCEL read write librarys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

import xlrd
import os



class myXLS:

    wb=0 #active workbook object
    ws=0 #active worksheet object

    isXLSX=False

    def __init__(self):
        ##print("...init myXLS")
        pass


    def openwb(self,filepath):
        #returns Workbook-object of openpyxl or xlrd

        #if is XLSX-Verison
        if filepath.split(".")[-1]=="xlsx":  # if .xlsx
            self.isXLSX=True
            ##print("---is XLSX")
            self.wb= load_workbook(filename=filepath,data_only=True,use_iterators=True)
        #if is XLS VErsion
        elif filepath.split(".")[-1]=="xls":  # if .xls
            self.isXLSX=False
            ##print("---is XLS")
            self.wb= xlrd.open_workbook(filepath)
        else:
            print("Wrong File Format! Cannot Open!")
            return False


    def my_get_sheet_names(self):  # return titles of sheets of active workbook
        if self.isXLSX:
            return self.wb.get_sheet_names()
        else:
            return self.wb.sheet_names()



    def my_activate_sheet(self,sheetname="-",index=-1):   # activates sheet (name)

        if sheetname!="-":
            if self.isXLSX:
                self.ws=self.wb.get_sheet_by_name(name=sheetname)
                ##print(" sheet activated: %s" %self.ws.title)
            else: #is xls
                self.ws=self.wb.sheet_by_name(sheetname)
                ##print(" sheet activated: %s" %self.ws.name)
        elif index!=-1:
            ##print("by index")
            if self.isXLSX:
                sheetname = self.my_get_sheet_names()[index] #sheet names of active workbook
                self.ws=self.wb.get_sheet_by_name(name=sheetname)
                ##print(" sheet activated: %s" %self.ws.title)
            else: #is xls
                self.ws=self.wb.sheet_by_index(sheetname)
                ##print(" sheet activated: %s" %self.ws.name)
        else:
            print("Open Worksheet - No correct Parameter given!")
            return False



    def my_cellvalue(self,cellcol=0,cellrow=0):  # start Column 0 Row 0 = A1
        if self.isXLSX:
            tmp=self._xy2xls(col=cellcol,row=cellrow) # 1,2 -> B2
            #print("found %s in %s" %(self.ws.cell(tmp).value,self.ws.title))
            return self.ws.cell(tmp).value
        else:
            return self.ws.cell_value(cellcol,cellrow)

    def wtf(self,cellcol=0,cellrow=0):
        tmp=self._xy2xls(col=cellcol,row=cellrow) # 1,2 -> B2
        print("made from col %s row %s : %s" %(cellcol,cellrow,tmp))
        print("wtf %s " %self.ws.cell(tmp).value)
        return True

    def my_findInArea(self,stext="-",area="A1:A2",get_content=False,first_only=False,get_allvalues=False):
        # no flag - return cells containing text ["A1","A2","A3"]
        # flag first_only ["A1"]
        # flag get_content -> ["textOfCell","textOfCell","textOfCell"]
        results=[]
        # if allvalues-flag is active - other flags are invalid
        if get_allvalues==True:
            get_content=False
            first_only=False

        #get search-area xy-coordinates
        cell_A= self._xls2xy(area.split(":")[0])          #"A2"-> [0,1] toupless
        cell_B= self._xls2xy(area.split(":")[1])         #"D8" -> [4,9]touples

        #swap smaller values into CellA(starting cell)
        for i in range(2):
            if cell_A[i]>cell_B[i]:
                cell_A[i],cell_B[i]=cell_B[i],cell_A[i]

##        print("starting cells")
##        print cell_A
##        print cell_B

        # iterate through area
        for j in range(cell_A[1],cell_B[1]+1):    #for each row  1-9
            if get_allvalues==True:
                results.append([])
##            print("j : %s" %j)
            for i in range(cell_A[0],cell_B[0]+1):        #for each column 0-4
                cellcontent=str(self.my_cellvalue(cellcol=i,cellrow=j))
##                print("i : %s" %i)

                if get_allvalues==True:
                    results[j-cell_A[1]].append(cellcontent)
                    continue

                if cellcontent.find(stext)!=-1:  #if substringis founc

                    if get_content==True:  # result append cell content
                        results.append(str(self.my_cellvalue(cellcol=i,cellrow=j)))
                    else:
                        results.append(self._xy2xls(col=i,row=j))

                    if first_only==True:
                        return results

        if results.__len__()==1 and get_allvalues==True: # to return [val,val] instead of [[val,val]]
            return results[0]   # if resutls are only oneline -> [[val,val]] the outer breaks are unnecessary

        return results




    def getRowRight(self,keycell="A1",number=1):  #returns content of cells right to key cell.

        pass


    def getColBelow(self,keycell="A1",number=1):  #returns content of cells below to key cell.
        pass



    def _xls2xy(self,xlscoord):  #  # "A8" -> "[0,9]
        cl=rw=""
        #....split
        for c in xlscoord.upper():
            if ord(c)<=90 and ord(c)>=65: ## is a char A-Z
                cl+=c
            if ord(c)<=57 and ord(c)>=48: ## is a value 0-9
                rw+=c

        rw=int(rw)-1                            #rw="1823" --> rw=1822
        cl=int(self._xcol2num(cl))              #cl="AA" -> 27

        return [cl,rw]                          # return [27,1824]
        #return [int(self._xcol2num(cl)),int(rw)-1]


    def _xy2xls(self,col=-1,row=-1,xy=-1):   # "col=0 row=9 ->"A10"  or [0,9]->"A10"
        if xy!=-1:  # col& row befuellt)
            col=xy[0]
            row=xy[1]

        if col==-1 and row==-1: # in this case no parameter were transmittet -> return false
            return False

        col=self._num2xcol(col)  # 0 -> "A"
        row+=1  # 9 -> 8

        return col+str(row)  # return "A8"




    def _xcol2num(self,coltxt):  # converts "AA" -> 26*1 + 0 = 26
        val=0
        for i,txt in enumerate(coltxt): #zb BCA
            val+= (ord(coltxt[coltxt.__len__()-1-i])-64) *(26**i)
        return val-1                                                        #-1 to make starting point =0


    def _num2xcol(self,col_nr):  # 0 -> "A"  , 8 -> "I"  126 ->"DW"
        #print("given : %s" % col_nr)
        col_nr=int(col_nr)
        txt=""
        for i in range(2,0,-1):   # 2 runden rest kommt separat drauf weil + 65 statt +64
            tmp=int(col_nr/(26**i))
            rest=int(col_nr%(26**i))
            if tmp<>0:
                txt+=chr(tmp+64)
            col_nr=rest

        return txt+chr(rest+65)


    def makeArea(self, col1=0, row1=0, col2=0, row2=0): #[1,1,2,2]-> "B2:C3"
        rc1=self._xy2xls(col=col1, row=row1)
        rc2=self._xy2xls(col=col2, row=row2)
        return str(rc1)+":"+str(rc2)

